from openpyxl import load_workbook
from business.models.excel_line import ExcelLine
from business.utils import rsetattr


def read_laboratory_excel(excel_url):
    wb = load_workbook(excel_url, read_only=True)
    rows = wb['Annonces'].iter_rows(2)
    builder = ExcelLineBuilder(rows, laboratory_excel_receipt).build()
    lines = builder.get_lines()
    return lines


laboratory_excel_receipt = {
        0: lambda line, value: rsetattr(line, 'sale_offer.rank', value),
        1: lambda line, value: rsetattr(line, 'sale_offer.product.name', value),
        2: lambda line, value: rsetattr(line, 'sale_offer.product.dci', value),
        3: lambda line, value: rsetattr(line, 'sale_offer.product.principal_barcode', value),
        4: lambda line, value: rsetattr(line, 'sale_offer.product.laboratory.name', value),
        5: lambda line, value: rsetattr(line, 'sale_offer.product.weight', value),
        6: lambda line, value: rsetattr(line, 'sale_offer.product.product_type.name', value),
        7: lambda line, value: rsetattr(line, 'sale_offer.product.vat.value', value),
        8: lambda line, value: rsetattr(line, 'sale_offer.product.unit_price', value),
        9: lambda line, value: rsetattr(line, 'sale_offer.description', value),
        10: lambda line, value: rsetattr(line, 'sale_offer.distribution', value),
        11: lambda line, value: rsetattr(line, 'sale_offer.distribution.discounted_price', value),
        12: lambda line, value: rsetattr(line, 'sale_offer.distribution.sold_by', value),
        13: lambda line, value: rsetattr(line, 'sale_offer.distribution.free_unit', value),
        14: lambda line, value: rsetattr(line, 'sale_offer.owner_id', value),
    }


class ExcelLineBuilder:
    def __init__(self, rows, receipt):
        self.excel_lines = []
        self.receipt = receipt
        self.rows = rows

    def __build_line(self, row):
        line = ExcelLine()
        for key, func in self.receipt.items():
            func(line, row[key].value)
        line.supervisor.identify_errors()
        return line

    def get_lines(self):
        return self.excel_lines

    def build(self):
        for row in self.rows:
            current_line = self.__build_line(row)
            if len(self.excel_lines) > 0:
                excel_lines = self.merge_duplicate(current_line)
                self.excel_lines = excel_lines
            else:
                self.excel_lines.append(current_line)
        return self

    def merge_duplicate(self, current_line):
        last_line = self.excel_lines[-1]
        if last_line.sale_offer.should_merge(current_line.sale_offer):
            last_line.sale_offer.merge(current_line.sale_offer)
            return self.excel_lines[:-1] + [last_line]
        return self.excel_lines[:-1] + [last_line, current_line]



