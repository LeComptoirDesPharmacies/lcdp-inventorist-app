import json
import logging
from collections import defaultdict
from api.consume.gen.factory.models.assembly_output_inner import AssemblyOutputInner
from typing import List

import tablib
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from api.consume.gen.factory.models.any_distribution_mode import AnyDistributionMode
from api.consume.gen.factory.models.any_factory import AnyFactory
from api.consume.gen.factory.models.assembly_creation_parameters import AssemblyCreationParameters
from api.consume.gen.factory.models.offer_planification_factory_all_of_records import OfferPlanificationFactoryAllOfRecords
from api.consume.gen.factory.models.product_upsert_factory_all_of_records import ProductUpsertFactoryAllOfRecords
from api.consume.gen.factory.models.sale_offer_upsert_factory_all_of_records import SaleOfferUpsertFactoryAllOfRecords
from business.constant import CHUNK_SIZE
from business.models.sale_offer import UNITARY_DISTRIBUTION, RANGE_DISTRIBUTION, QUOTATION_DISTRIBUTION
from business.services.product import get_product_type_by_name
from business.services.providers import get_manage_assembly_api
from business.services.security import get_api_key
from business.services.user import get_current_user_id
from business.services.vat import get_vat_by_value
from business.utils import rgetattr, execution_time

def __prefetch(prefetch_type, keys_from_file, get_from_api, get_keys_from_api_object):
    packets = [list(keys_from_file)[i:i + CHUNK_SIZE] for i in range(0, len(keys_from_file), CHUNK_SIZE)]

    array_from_api = list(map(get_from_api, packets))

    flatten_array = []
    for sub_list in array_from_api:
        flatten_array.extend(sub_list)

    map_of_objects = {}
    for obj in flatten_array:
        keys = get_keys_from_api_object(obj)
        for key in keys:
            map_of_objects[key] = obj

    logging.info(f"Map of {prefetch_type} has {len(map_of_objects)} element(s)")

    return map_of_objects


def __build_product_upsert(product_line):
    product_type = get_product_type_by_name(product_line.product_type.name)
    vat = get_vat_by_value(product_line.vat.value)

    product_upsert = dict()

    if product_line.principal_barcode:
        product_upsert['principal_barcode'] = product_line.principal_barcode

    if product_line.name:
        product_upsert['name'] = product_line.name

    if product_line.dci:
        product_upsert['dci'] = product_line.dci

    if product_line.laboratory.name:
        product_upsert['laboratory_name'] = product_line.laboratory.name

    if product_line.weight:
        print("product_line.weight", product_line.weight)
        product_upsert['unit_weight'] = product_line.weight

    if product_line.unit_price:
        product_upsert['unit_price'] = product_line.unit_price

    if product_line.status:
        product_upsert['status'] = product_line.status.upper()

    if product_type:
        product_upsert['type_id'] = product_type.id

    if vat:
        product_upsert['vat_id'] = vat.id

    return product_upsert


def __build_distribution_mode(sale_offer_line):
    if not sale_offer_line.sale_offer.distribution or sale_offer_line.sale_offer.distribution.is_empty():
        return dict()

    if sale_offer_line.sale_offer.distribution_type == RANGE_DISTRIBUTION:
        ranges = []
        for range in sale_offer_line.sale_offer.distribution.ranges:
            new_range = {
                'quantity': range.sold_by,
                'unitPrice': range.discounted_price,
                'freeUnits': range.free_unit or 0
            }
            ranges.append(new_range)
        distribution_mode = {
            'type': 'RANGE',
            'ranges': ranges,
            'minimalQuantity': 1,
            'maximalQuantity': sale_offer_line.sale_offer.distribution.maximal_quantity
        }
    elif sale_offer_line.sale_offer.distribution_type == UNITARY_DISTRIBUTION:
        distribution_mode = {
            'type': 'UNITARY',
            'unitPrice': sale_offer_line.sale_offer.distribution.discounted_price,
            'soldBy': sale_offer_line.sale_offer.distribution.sold_by,
            'minimalQuantity': 1,
            'maximalQuantity': sale_offer_line.sale_offer.distribution.maximal_quantity
        }
    elif sale_offer_line.sale_offer.distribution_type == QUOTATION_DISTRIBUTION:
        distribution_mode = {
            'type': 'QUOTATION',
            'soldBy': sale_offer_line.sale_offer.distribution.sold_by,
            'minimalQuantity': 1,
            'maximalQuantity': sale_offer_line.sale_offer.distribution.maximal_quantity
        }
    else:
        distribution_mode = dict()

    return distribution_mode

def __build_stock(stock_line, full=False):
    '''
    :param stock_line:
    :param full: If'true', then all stock object will be built
        when at least one field is set. Useful for offer planification as row are complete
        but can be problematic when updating sale offers.
        TODO : We can not reset lapsing_date NULL and batch NULL with SALE_OFFER_UPSERT
    :return:
    '''
    stock = dict()

    if stock_line.remaining_quantity != None or full:
        remaining_quantity = stock_line.remaining_quantity
        try:
            int(remaining_quantity)
        except ValueError:
            # stock can contain non-breaking space character to separate thousands
            remaining_quantity = int(remaining_quantity.replace('\u202f', ''))
            pass
        stock['remaining_quantity'] = remaining_quantity

    if stock_line.lapsing_date != None or full:
        stock['lapsing_date'] = stock_line.lapsing_date

    if stock_line.batch != None or full:
        stock['batch'] = stock_line.batch

    return stock


@execution_time
def sale_offer_upsert_from_excel_lines(lines, filename, clean=False, **kwargs):
    logging.info(f"{len(lines)} excel line(s) are candide for sale offer modification/creation")

    items_by_owner = defaultdict(list)

    for line in lines:
        product_upsert = __build_product_upsert(line.sale_offer.product)
        distribution_mode = __build_distribution_mode(line)
        stock = __build_stock(line.sale_offer.stock)

        new_item = dict()

        new_item['reference'] = line.sale_offer.reference

        if product_upsert:
            new_item['product'] = product_upsert

        if distribution_mode:
            new_item['distributionMode'] = AnyDistributionMode.from_dict(distribution_mode)

        if stock:
            new_item['stock'] = stock

        if line.sale_offer.status:
            new_item['status'] = line.sale_offer.status.upper()

        if line.sale_offer.rank:
            new_item['rank'] = line.sale_offer.rank

        if line.sale_offer.description:
            new_item['description'] = line.sale_offer.description

        owner_id = line.sale_offer.owner_id

        if owner_id:
            items_by_owner[owner_id].append(
                SaleOfferUpsertFactoryAllOfRecords(**new_item)
            )

    manage_assembly_api = get_manage_assembly_api()

    for owner_id, items in items_by_owner.items():
        manage_assembly_api.create_assembly(
            AssemblyCreationParameters(
                owner_id=get_current_user_id(),
                tags=['seller-id:{}'.format(owner_id), 'filename:{}'.format(filename)],
                factory=AnyFactory({
                    'type': 'SALE_OFFER_UPSERT',
                    'clean': clean,
                    'sellerId': owner_id,
                    'records': items
                })),
            _request_auth=manage_assembly_api.api_client.create_auth_settings("apiKeyAuth", get_api_key()),
        )


@execution_time
def create_offer_planificiation_from_excel_lines(lines, filename, clean=False, **kwargs):
    logging.info(f"{len(lines)} excel line(s) are candide for sale offer modification/creation")

    items_by_owner = defaultdict(list)

    for line in lines:
        product_upsert = __build_product_upsert(line.sale_offer.product)
        distribution_mode = __build_distribution_mode(line)

        stock = __build_stock(line.sale_offer.stock, full=True)

        new_item = dict()

        if product_upsert:
            new_item['product'] = product_upsert

        if distribution_mode:
            new_item['distributionMode'] = AnyDistributionMode.from_dict(distribution_mode)

        if stock:
            new_item['stock'] = stock

        if line.sale_offer.status:
            new_item['status'] = line.sale_offer.status.upper()

        if line.sale_offer.rank:
            new_item['rank'] = line.sale_offer.rank

        if line.sale_offer.description:
            new_item['description'] = line.sale_offer.description

        owner_id = line.sale_offer.owner_id

        if owner_id:
            items_by_owner[owner_id].append(
                OfferPlanificationFactoryAllOfRecords(**new_item)
            )

    manage_assembly_api = get_manage_assembly_api()

    for owner_id, items in items_by_owner.items():
        manage_assembly_api.create_assembly(
            AssemblyCreationParameters(
                owner_id=get_current_user_id(),
                tags=['seller-id:{}'.format(owner_id), 'filename:{}'.format(filename)],
                factory=AnyFactory({
                    'type': 'OFFER_PLANIFICATION',
                    'clean': clean,
                    'offerName': None, # Update every sale offer, without taking care of the offer name
                    'sellerId': owner_id,
                    'records': items
                })),
            _request_auth=manage_assembly_api.api_client.create_auth_settings("apiKeyAuth", get_api_key()),
        )


@execution_time
def product_upsert_from_excel_lines(lines, filename, **kwargs):
    logging.info(f"{len(lines)} excel line(s) are candide for product modification/creation")

    items = []

    for line in lines:
        product_upsert = __build_product_upsert(line.sale_offer.product)

        new_item = dict()

        if product_upsert:
            new_item['product'] = product_upsert

        items.append(ProductUpsertFactoryAllOfRecords(**new_item))

    manage_assembly_api = get_manage_assembly_api()

    manage_assembly_api.create_assembly(
        AssemblyCreationParameters(
            owner_id=get_current_user_id(),
            tags=['filename:{}'.format(filename)],
            factory=AnyFactory({
                'type': 'PRODUCT_UPSERT',
                'records': items,
            })
        ),
        _request_auth=manage_assembly_api.api_client.create_auth_settings("apiKeyAuth", get_api_key()),
    )


def excel_to_dict(obj_class, excel_path, excel_mapper, sheet_name, header_row,
                  min_row, max_row=None, obj_unique_key=None, custom_dict=None):
    results = {}
    if isinstance(custom_dict, dict):
        results = custom_dict

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]
        column_indices = {col: cell.value for col, cell in enumerate(ws[header_row])}
        max_col = len(column_indices.keys())
        for idx, row in enumerate(ws.iter_rows(min_row=min_row, max_row=max_row, max_col=max_col, values_only=True)):
            # check if row is empty
            if not any(row):
                continue

            obj = obj_class()

            cells = {
                column_indices[col]: int(value) if isinstance(value, float) and value.is_integer() else value
                for col, value in enumerate(row)
                if value is not None
            }

            for col in excel_mapper:
                col.set_from_excel(obj, cells.get(col.excel_column_name, None))
            if obj_unique_key:
                obj_key = rgetattr(obj, obj_unique_key, None)
                results[obj_key] = obj
            else:
                results[idx] = obj
    finally:
        wb.close()

    return results


def dict_to_excel(assembly_output_list: List[AssemblyOutputInner], excel_path):
    dataset = tablib.Dataset()

    # sort headers by name
    headers = list()
    for header in sorted(assembly_output_list[0].unit.keys()):
        headers.append(header)

    dataset.headers = headers
    dataset.headers.append('status_comment')

    # build dataset
    for assembly_output in assembly_output_list:
        unit_value = []
        for header in headers:
            header_value = assembly_output.unit.get(header, None)
            if isinstance(header_value, dict):
                header_value = json.dumps(header_value, indent=2)
            unit_value.append(header_value)
        unit_value.append(assembly_output.status_comment)

        dataset.rpush(unit_value, tags=[assembly_output.status])

    tabSuccess = "Succès"
    tabErrors = "Erreurs"

    # Créer un Dataset pour les données "succeeded"
    dataset_succeeded = dataset.filter('SUCCEEDED')
    dataset_succeeded.title = tabSuccess
    # Créer un Dataset pour les données "failed"
    dataset_failed = dataset.filter('FAILED')
    dataset_failed.title = tabErrors

    workbook = tablib.Databook()
    workbook.add_sheet(dataset_succeeded)
    workbook.add_sheet(dataset_failed)

    with open(excel_path, 'wb') as f:
        f.write(workbook.xlsx)

    # Resize columns width for readability
    wb = load_workbook(excel_path)
    resize_column_width(wb.get_sheet_by_name(tabSuccess))
    resize_column_width(wb.get_sheet_by_name(tabErrors))
    wb.save(excel_path)

    return excel_path

def resize_column_width(ws: Worksheet):
    for column in ws.columns:
        column_letter = get_column_letter(column[0].column)
        # set arbitrary to 50
        ws.column_dimensions[column_letter].width = 50


def clean_sale_offers(lines):
    print("----------------------")
    print("TODO clean_sale_offers")
    print("----------------------")
